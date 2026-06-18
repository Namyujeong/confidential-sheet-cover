from __future__ import annotations

import argparse
import csv
import hashlib
import json
import shutil
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_KEYWORDS = (
    # English
    "payroll",
    "salary",
    "salaries",
    "wage",
    "wages",
    "compensation",
    "bonus",
    "severance",
    "retirement",
    "pension",
    "benefits",
    "personnel cost",
    "labor cost",
    "labour cost",
    "headcount cost",
    "project cost",
    "business cost",
    "business expense",
    "operating expense",
    "grant budget",
    "cost allocation",
    # Korean
    "인건비",
    "연봉",
    "임금",
    "급여",
    "보수",
    "보상",
    "상여",
    "상여금",
    "퇴직금",
    "퇴직정산",
    "포괄임금",
    "사업비",
    "운영비",
    "예산",
)

DEFAULT_COVER_SHEET_NAME = "Confidential"
DEFAULT_COVER_TEXT = "Confidential Document - details start on the next sheet"
DEFAULT_OUTPUT_DIR = "confidential-cover-results"


@dataclass
class ScanResult:
    matched_keywords: set[str]
    matched_locations: list[str]
    first_sheet: str
    error: str = ""


def normalize_text(value: object) -> str:
    return unicodedata.normalize("NFKC", str(value)).casefold()


def find_keyword_matches(text: object, keywords: Iterable[str]) -> list[str]:
    normalized = normalize_text(text)
    return [keyword for keyword in keywords if normalize_text(keyword) in normalized]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_keywords(args: argparse.Namespace) -> list[str]:
    keywords: list[str] = []
    if not args.no_default_keywords:
        keywords.extend(DEFAULT_KEYWORDS)
    for keyword in args.keyword or []:
        keyword = keyword.strip()
        if keyword:
            keywords.append(keyword)
    if args.keywords_file:
        with Path(args.keywords_file).open(encoding="utf-8") as handle:
            for line in handle:
                keyword = line.strip()
                if keyword and not keyword.startswith("#"):
                    keywords.append(keyword)

    seen: set[str] = set()
    unique: list[str] = []
    for keyword in keywords:
        key = normalize_text(keyword)
        if key not in seen:
            seen.add(key)
            unique.append(keyword)
    return unique


def iter_workbooks(roots: list[Path]) -> Iterable[Path]:
    seen: set[Path] = set()
    for root in roots:
        if root.is_file():
            candidates = [root]
        else:
            candidates = root.rglob("*.xlsx")
        for path in candidates:
            if path.name.startswith("~$") or path.suffix.lower() != ".xlsx":
                continue
            resolved = path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            yield path


def display_path(path: Path, roots: list[Path]) -> str:
    for root in roots:
        base = root if root.is_dir() else root.parent
        try:
            return str(path.relative_to(base))
        except ValueError:
            continue
    return str(path)


def add_match(
    text: object,
    location: str,
    keywords: list[str],
    matched_keywords: set[str],
    matched_locations: list[str],
    max_locations: int,
) -> None:
    matches = find_keyword_matches(text, keywords)
    for keyword in matches:
        matched_keywords.add(keyword)
    if matches and len(matched_locations) < max_locations:
        matched_locations.append(f"{location}: {', '.join(matches)}")


def scan_workbook(
    path: Path,
    roots: list[Path],
    keywords: list[str],
    scan_content: bool,
    max_cells_per_sheet: int,
    max_locations: int,
) -> ScanResult:
    matched_keywords: set[str] = set()
    matched_locations: list[str] = []
    add_match(display_path(path, roots), "path", keywords, matched_keywords, matched_locations, max_locations)

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return ScanResult(matched_keywords, matched_locations, "", f"{type(exc).__name__}: {exc}")

    first_sheet = workbook.sheetnames[0] if workbook.sheetnames else ""
    for sheet_name in workbook.sheetnames:
        add_match(sheet_name, f"sheet:{sheet_name}", keywords, matched_keywords, matched_locations, max_locations)

    if scan_content:
        for worksheet in workbook.worksheets:
            scanned = 0
            for row in worksheet.iter_rows():
                for cell in row:
                    scanned += 1
                    if isinstance(cell.value, str):
                        add_match(
                            cell.value,
                            f"{worksheet.title}!{cell.coordinate}",
                            keywords,
                            matched_keywords,
                            matched_locations,
                            max_locations,
                        )
                    if scanned >= max_cells_per_sheet:
                        break
                if scanned >= max_cells_per_sheet:
                    break

    workbook.close()
    return ScanResult(matched_keywords, matched_locations, first_sheet)


def unique_backup_path(path: Path, roots: list[Path], backup_dir: Path) -> Path:
    relative = Path(display_path(path, roots))
    fingerprint = hashlib.sha1(str(path.resolve()).encode("utf-8", "surrogateescape")).hexdigest()[:10]
    return backup_dir / relative.parent / f"{path.stem}.{fingerprint}.before-confidential{path.suffix}"


def unique_sheet_name(workbook, preferred: str) -> str:
    if preferred not in workbook.sheetnames:
        return preferred
    index = 2
    while f"{preferred} {index}" in workbook.sheetnames:
        index += 1
    return f"{preferred} {index}"


def style_cover_sheet(worksheet, cover_text: str) -> None:
    worksheet["A1"] = cover_text
    worksheet.merge_cells("A1:F3")
    worksheet["A1"].font = Font(bold=True, size=16, color="1F2937")
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet["A1"].fill = PatternFill(fill_type="solid", fgColor="F3F4F6")
    for column in "ABCDEF":
        worksheet.column_dimensions[column].width = 18
    for row in range(1, 4):
        worksheet.row_dimensions[row].height = 28
    worksheet.sheet_view.showGridLines = False


def apply_cover(
    path: Path,
    roots: list[Path],
    backup_dir: Path,
    sheet_name: str,
    cover_text: str,
) -> dict[str, str]:
    before_hash = sha256(path)
    workbook = load_workbook(path)
    first_before = workbook.sheetnames[0] if workbook.sheetnames else ""

    if first_before == sheet_name and workbook[sheet_name]["A1"].value == cover_text:
        workbook.close()
        return {
            "status": "skipped_already_covered",
            "first_sheet_before": first_before,
            "first_sheet_after": first_before,
            "backup_path": "",
            "sha256_before": before_hash,
            "sha256_after": before_hash,
        }

    backup_path = unique_backup_path(path, roots, backup_dir)
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(path, backup_path)

    if sheet_name in workbook.sheetnames and workbook[sheet_name]["A1"].value == cover_text:
        worksheet = workbook[sheet_name]
        offset = -workbook.index(worksheet)
        if offset:
            workbook.move_sheet(worksheet, offset=offset)
        status = "moved_existing_cover"
    else:
        worksheet = workbook.create_sheet(unique_sheet_name(workbook, sheet_name), 0)
        style_cover_sheet(worksheet, cover_text)
        status = "updated"

    workbook.active = 0
    workbook.save(path)
    workbook.close()

    verify = load_workbook(path, read_only=True, data_only=False)
    first_after = verify.sheetnames[0] if verify.sheetnames else ""
    cover_ok = bool(first_after and first_after.startswith(sheet_name) and verify[first_after]["A1"].value == cover_text)
    verify.close()

    return {
        "status": status if cover_ok else "error_verify_failed",
        "first_sheet_before": first_before,
        "first_sheet_after": first_after,
        "backup_path": str(backup_path),
        "sha256_before": before_hash,
        "sha256_after": sha256(path),
    }


def process_file(path: Path, roots: list[Path], args: argparse.Namespace, keywords: list[str]) -> dict[str, str]:
    row = {
        "path": str(path),
        "relative_path": display_path(path, roots),
        "status": "",
        "matched_keywords": "",
        "matched_locations": "",
        "first_sheet_before": "",
        "first_sheet_after": "",
        "backup_path": "",
        "sha256_before": "",
        "sha256_after": "",
        "error": "",
    }

    scan = scan_workbook(
        path,
        roots,
        keywords,
        scan_content=not args.no_content_scan,
        max_cells_per_sheet=args.max_cells_per_sheet,
        max_locations=args.max_match_locations,
    )
    row["matched_keywords"] = "|".join(sorted(scan.matched_keywords, key=normalize_text))
    row["matched_locations"] = " | ".join(scan.matched_locations)
    row["first_sheet_before"] = scan.first_sheet

    if scan.error and not scan.matched_keywords:
        row["status"] = "error"
        row["error"] = scan.error
        return row

    if not scan.matched_keywords:
        row["status"] = "skipped_no_match"
        return row

    if args.dry_run:
        row["status"] = "dry_run_target"
        row["first_sheet_after"] = args.sheet_name
        if scan.error:
            row["error"] = f"matched but workbook scan warning: {scan.error}"
        return row

    try:
        result = apply_cover(path, roots, Path(args.backup_dir), args.sheet_name, args.cover_text)
    except Exception as exc:
        row["status"] = "error"
        row["error"] = f"{type(exc).__name__}: {exc}"
        return row

    row.update(result)
    if scan.error:
        row["error"] = f"matched but workbook scan warning: {scan.error}"
    return row


def write_outputs(rows: list[dict[str, str]], output_dir: Path, dry_run: bool, keywords: list[str]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    fields = [
        "path",
        "relative_path",
        "status",
        "matched_keywords",
        "matched_locations",
        "first_sheet_before",
        "first_sheet_after",
        "backup_path",
        "sha256_before",
        "sha256_after",
        "error",
    ]
    with (output_dir / "manifest.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)

    counts: dict[str, int] = {}
    for row in rows:
        counts[row["status"]] = counts.get(row["status"], 0) + 1
    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "dry_run": dry_run,
        "file_count": len(rows),
        "status_counts": counts,
        "keyword_count": len(keywords),
        "outputs": {
            "manifest": str(output_dir / "manifest.csv"),
            "summary": str(output_dir / "summary.json"),
        },
    }
    (output_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Find sensitive .xlsx workbooks and add a confidential cover sheet as the first worksheet.",
    )
    parser.add_argument("roots", nargs="+", help="Files or folders to scan. Google Drive for Desktop folders work as normal folders.")
    parser.add_argument("--dry-run", action="store_true", help="Only report matching files; do not modify workbooks.")
    parser.add_argument("--limit", type=int, default=0, help="Process only the first N .xlsx files after sorting.")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="Directory for manifest.csv and summary.json.")
    parser.add_argument("--backup-dir", help="Directory for backups. Defaults to <output-dir>/backups.")
    parser.add_argument("--sheet-name", default=DEFAULT_COVER_SHEET_NAME, help="Name for the confidential cover sheet.")
    parser.add_argument("--cover-text", default=DEFAULT_COVER_TEXT, help="Text to place in A1 of the cover sheet.")
    parser.add_argument("--keyword", action="append", help="Additional keyword to match. Can be repeated.")
    parser.add_argument("--keywords-file", help="UTF-8 text file with one keyword per line.")
    parser.add_argument("--no-default-keywords", action="store_true", help="Use only --keyword and --keywords-file values.")
    parser.add_argument("--no-content-scan", action="store_true", help="Scan file paths and sheet names only, not cell text.")
    parser.add_argument("--max-cells-per-sheet", type=int, default=5000, help="Maximum cells to inspect per sheet.")
    parser.add_argument("--max-match-locations", type=int, default=25, help="Maximum match locations recorded per workbook.")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    roots = [Path(root).expanduser().resolve() for root in args.roots]
    missing = [str(root) for root in roots if not root.exists()]
    if missing:
        parser.error(f"path does not exist: {', '.join(missing)}")

    output_dir = Path(args.output_dir).expanduser().resolve()
    args.output_dir = str(output_dir)
    if not args.backup_dir:
        args.backup_dir = str(output_dir / "backups")

    keywords = load_keywords(args)
    if not keywords:
        parser.error("no keywords configured")

    targets = sorted(iter_workbooks(roots), key=lambda item: display_path(item, roots))
    if args.limit:
        targets = targets[: args.limit]

    rows: list[dict[str, str]] = []
    for index, path in enumerate(targets, 1):
        row = process_file(path, roots, args, keywords)
        rows.append(row)
        print(f"{index}/{len(targets)} {row['status']} {row['relative_path']}", flush=True)

    write_outputs(rows, output_dir, args.dry_run, keywords)
    print(f"wrote {output_dir / 'manifest.csv'}", file=sys.stderr)
    print(f"wrote {output_dir / 'summary.json'}", file=sys.stderr)
    return 1 if any(row["status"].startswith("error") for row in rows) else 0
